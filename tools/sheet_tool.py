from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
import re
from datetime import datetime

usernames = {'Ben-Kinkor':1, 'Colin-Recker':2, 'Fineas-Herrera':3, 'jj-timo':4, 'Lucas-Johnson':5, 'Quinn-Carlson':6, 'Noah-Lau':7}
wb = load_workbook('RooftopLedger.xlsx')
ws = wb.active

def findMonthCell(month):
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        for cell in row:
            print(cell)

findMonthCell('may')

#Finds and returns the correct cell that corresponds to a given name, month, and cell type
def findEntryCell(username, monthcell, type):
    monthCol = re.findall("\D", monthcell)
    lastCol = ord(monthCol[-1])
    monthRow =  "".join(re.findall("\d", monthcell))
    targetCol = "".join(monthCol[:-1])+chr(lastCol+2*type+1)
    targetRow = int(monthRow) + usernames[username] + 1
    cell = str(targetCol) + str(targetRow)
    print(cell)
    return cell

#Updates a specific entry in the spreadsheet
def updateCellDate(cell, date) -> None:
    targetCell = ws[cell]
    targetCell.value = date
    targetCell.number_format = 'DD/MM/YY'
    wb.save('RooftopLedger.xlsx')
    print(f"Updated {cell} to {date}")

def updateAll(month):
    return

#user, month cell, type
user, month, type, date = findTransaction()

targetCell = findEntryCell('Ben-Kinkor', "I50", 1)
updateCellDate(targetCell, date)

